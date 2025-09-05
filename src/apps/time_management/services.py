"""
OneSquare Time Management - 비즈니스 로직 서비스

업무시간 관리의 핵심 비즈니스 로직을 담당하는 서비스 레이어
"""

from django.db import transaction
from django.utils import timezone
from django.db.models import Sum, Count, Avg, Q
from django.contrib.auth import get_user_model
from datetime import datetime, timedelta, date, time
from typing import Dict, List, Optional, Tuple
import logging
from decimal import Decimal

from .models import (
    WorkTimeRecord, 
    WorkTimeSettings, 
    WorkTimeSummary, 
    OverTimeRule
)
from apps.notion_api.services import NotionSyncService

User = get_user_model()
logger = logging.getLogger(__name__)


class WorkTimeTrackingService:
    """
    근무시간 추적 핵심 서비스
    """
    
    def __init__(self):
        self.notion_service = NotionSyncService()
    
    def check_in_user(self, user: User, location_data: Dict = None) -> WorkTimeRecord:
        """
        사용자 출근 처리
        
        Args:
            user: 출근할 사용자
            location_data: GPS 위치 정보 (선택적)
            
        Returns:
            WorkTimeRecord: 생성된 근무시간 기록
        """
        today = timezone.now().date()
        
        # 오늘 이미 출근한 기록이 있는지 확인
        existing_record = WorkTimeRecord.objects.filter(
            user=user,
            work_date=today
        ).first()
        
        if existing_record and existing_record.check_in_time:
            raise ValueError(f"오늘 이미 출근 처리되었습니다. (출근시간: {existing_record.check_in_time.strftime('%H:%M')})")
        
        try:
            with transaction.atomic():
                if existing_record:
                    # 기존 기록에 출근시간 추가
                    existing_record.check_in_time = timezone.now()
                    existing_record.check_in_location = location_data or {}
                    existing_record.status = WorkTimeRecord.RecordStatus.CONFIRMED
                    existing_record.save()
                    record = existing_record
                else:
                    # 새로운 기록 생성
                    record = WorkTimeRecord.objects.create(
                        user=user,
                        work_date=today,
                        check_in_time=timezone.now(),
                        check_in_location=location_data or {},
                        status=WorkTimeRecord.RecordStatus.CONFIRMED
                    )
                
                logger.info(f"출근 처리 완료: {user.username} at {record.check_in_time}")
                return record
                
        except Exception as e:
            logger.error(f"출근 처리 실패: {user.username} - {str(e)}")
            raise
    
    def check_out_user(self, user: User, location_data: Dict = None) -> WorkTimeRecord:
        """
        사용자 퇴근 처리
        
        Args:
            user: 퇴근할 사용자
            location_data: GPS 위치 정보 (선택적)
            
        Returns:
            WorkTimeRecord: 업데이트된 근무시간 기록
        """
        today = timezone.now().date()
        
        # 오늘의 출근 기록 찾기
        record = WorkTimeRecord.objects.filter(
            user=user,
            work_date=today,
            check_in_time__isnull=False
        ).first()
        
        if not record:
            raise ValueError("출근 기록이 없습니다. 먼저 출근 처리를 해주세요.")
        
        if record.check_out_time:
            raise ValueError(f"이미 퇴근 처리되었습니다. (퇴근시간: {record.check_out_time.strftime('%H:%M')})")
        
        try:
            with transaction.atomic():
                record.check_out_time = timezone.now()
                record.check_out_location = location_data or {}
                record.save()  # save() 시 자동으로 근무시간 계산됨
                
                # 초과근무 규정 확인 및 알림
                if record.is_overtime:
                    self._handle_overtime_notification(record)
                
                logger.info(f"퇴근 처리 완료: {user.username} - 근무시간: {record.get_work_time_formatted()}")
                return record
                
        except Exception as e:
            logger.error(f"퇴근 처리 실패: {user.username} - {str(e)}")
            raise
    
    def get_today_work_status(self, user: User) -> Dict:
        """
        오늘의 근무 상태 조회
        
        Args:
            user: 조회할 사용자
            
        Returns:
            Dict: 근무 상태 정보
        """
        today = timezone.now().date()
        
        record = WorkTimeRecord.objects.filter(
            user=user,
            work_date=today
        ).first()
        
        status = {
            'date': today,
            'is_checked_in': False,
            'is_checked_out': False,
            'check_in_time': None,
            'check_out_time': None,
            'current_work_time': 0,
            'status': 'not_started'
        }
        
        if record:
            status.update({
                'is_checked_in': bool(record.check_in_time),
                'is_checked_out': bool(record.check_out_time),
                'check_in_time': record.check_in_time,
                'check_out_time': record.check_out_time,
            })
            
            if record.check_in_time and record.check_out_time:
                status['status'] = 'completed'
                status['current_work_time'] = record.actual_work_minutes
            elif record.check_in_time:
                status['status'] = 'in_progress'
                # 현재까지의 근무시간 계산
                current_minutes = int((timezone.now() - record.check_in_time).total_seconds() / 60)
                status['current_work_time'] = current_minutes
        
        return status
    
    def adjust_work_time(self, record_id: int, check_in_time: datetime, 
                        check_out_time: datetime, reason: str, 
                        adjusted_by: User) -> WorkTimeRecord:
        """
        근무시간 수동 조정
        
        Args:
            record_id: 조정할 기록 ID
            check_in_time: 수정할 출근시간
            check_out_time: 수정할 퇴근시간
            reason: 조정 사유
            adjusted_by: 조정자
            
        Returns:
            WorkTimeRecord: 조정된 기록
        """
        try:
            with transaction.atomic():
                record = WorkTimeRecord.objects.select_for_update().get(id=record_id)
                
                # 원본 데이터 백업
                original_data = {
                    'check_in_time': record.check_in_time,
                    'check_out_time': record.check_out_time,
                    'total_work_minutes': record.total_work_minutes,
                    'actual_work_minutes': record.actual_work_minutes,
                }
                
                # 새로운 시간 설정
                record.check_in_time = check_in_time
                record.check_out_time = check_out_time
                record.adjustment_reason = reason
                record.status = WorkTimeRecord.RecordStatus.MODIFIED
                record.approved_by = adjusted_by
                record.approved_at = timezone.now()
                
                record.save()  # 자동 계산 수행
                
                logger.info(
                    f"근무시간 조정 완료: {record.user.username} - "
                    f"{original_data['total_work_minutes']}분 -> {record.total_work_minutes}분 "
                    f"(조정자: {adjusted_by.username})"
                )
                
                return record
                
        except WorkTimeRecord.DoesNotExist:
            raise ValueError("해당 근무 기록을 찾을 수 없습니다.")
        except Exception as e:
            logger.error(f"근무시간 조정 실패: record_id={record_id} - {str(e)}")
            raise
    
    def _handle_overtime_notification(self, record: WorkTimeRecord):
        """초과근무 알림 처리"""
        # TODO: 실제 알림 시스템 연동
        logger.info(
            f"초과근무 발생: {record.user.username} - "
            f"{record.overtime_hours:.1f}시간 초과"
        )
    
    def sync_to_notion(self, record: WorkTimeRecord) -> bool:
        """
        Notion과 동기화
        
        Args:
            record: 동기화할 근무시간 기록
            
        Returns:
            bool: 동기화 성공 여부
        """
        try:
            properties = record.to_notion_properties()
            
            if record.notion_page_id:
                # 기존 페이지 업데이트
                result = self.notion_service.update_page(
                    record.notion_page_id,
                    properties
                )
            else:
                # 새 페이지 생성
                result = self.notion_service.create_page(
                    properties
                )
                record.notion_page_id = result.get('id', '')
            
            record.is_notion_synced = True
            record.notion_last_synced = timezone.now()
            record.save(update_fields=['notion_page_id', 'is_notion_synced', 'notion_last_synced'])
            
            logger.info(f"Notion 동기화 완료: {record.user.username} - {record.work_date}")
            return True
            
        except Exception as e:
            logger.error(f"Notion 동기화 실패: {record.user.username} - {record.work_date} - {str(e)}")
            return False


class WorkTimeStatisticsService:
    """
    근무시간 통계 서비스
    """
    
    def calculate_daily_summary(self, user: User, target_date: date) -> Dict:
        """
        일간 근무시간 요약 계산
        
        Args:
            user: 대상 사용자
            target_date: 대상 날짜
            
        Returns:
            Dict: 일간 통계 데이터
        """
        record = WorkTimeRecord.objects.filter(
            user=user,
            work_date=target_date
        ).first()
        
        if not record:
            return {
                'date': target_date,
                'work_status': 'absent',
                'total_work_minutes': 0,
                'actual_work_minutes': 0,
                'overtime_minutes': 0,
                'is_late': False,
                'is_early_leave': False,
            }
        
        # 표준 근무시간 (9시-18시) 기준으로 지각/조퇴 판단
        standard_start = time(9, 0)
        standard_end = time(18, 0)
        
        is_late = (record.check_in_time and 
                  record.check_in_time.time() > standard_start)
        is_early_leave = (record.check_out_time and 
                         record.check_out_time.time() < standard_end and
                         record.actual_work_minutes < 480)  # 8시간 미만
        
        return {
            'date': target_date,
            'record_id': record.id,
            'work_status': 'present' if record.actual_work_minutes > 0 else 'absent',
            'check_in_time': record.check_in_time,
            'check_out_time': record.check_out_time,
            'total_work_minutes': record.total_work_minutes,
            'actual_work_minutes': record.actual_work_minutes,
            'overtime_minutes': record.overtime_minutes,
            'is_overtime': record.is_overtime,
            'is_undertime': record.is_undertime,
            'is_late': is_late,
            'is_early_leave': is_early_leave,
            'work_time_formatted': record.get_work_time_formatted(),
        }
    
    def calculate_weekly_summary(self, user: User, year: int, week: int) -> Dict:
        """
        주간 근무시간 요약 계산
        
        Args:
            user: 대상 사용자
            year: 연도
            week: 주차
            
        Returns:
            Dict: 주간 통계 데이터
        """
        # 해당 주의 시작일과 종료일 계산
        start_date = datetime.strptime(f'{year}-W{week:02d}-1', "%Y-W%U-%w").date()
        end_date = start_date + timedelta(days=6)
        
        records = WorkTimeRecord.objects.filter(
            user=user,
            work_date__range=[start_date, end_date]
        ).aggregate(
            total_work=Sum('total_work_minutes'),
            actual_work=Sum('actual_work_minutes'),
            overtime=Sum('overtime_minutes'),
            work_days=Count('id', filter=Q(actual_work_minutes__gt=0))
        )
        
        # 지각/조퇴 계산
        standard_start = time(9, 0)
        standard_end = time(18, 0)
        
        late_count = WorkTimeRecord.objects.filter(
            user=user,
            work_date__range=[start_date, end_date],
            check_in_time__time__gt=standard_start
        ).count()
        
        early_leave_count = WorkTimeRecord.objects.filter(
            user=user,
            work_date__range=[start_date, end_date],
            check_out_time__time__lt=standard_end,
            actual_work_minutes__lt=480
        ).count()
        
        total_work_minutes = records['total_work'] or 0
        actual_work_minutes = records['actual_work'] or 0
        overtime_minutes = records['overtime'] or 0
        work_days = records['work_days'] or 0
        
        # 주간 표준 근무시간 (5일 × 8시간)
        weekly_standard_minutes = 2400
        
        return {
            'year': year,
            'week': week,
            'period': f"{start_date} ~ {end_date}",
            'total_work_minutes': total_work_minutes,
            'actual_work_minutes': actual_work_minutes,
            'overtime_minutes': overtime_minutes,
            'work_days': work_days,
            'expected_work_days': 5,  # 주 5일 근무 기준
            'late_count': late_count,
            'early_leave_count': early_leave_count,
            'total_work_hours': round(total_work_minutes / 60, 2),
            'actual_work_hours': round(actual_work_minutes / 60, 2),
            'overtime_hours': round(overtime_minutes / 60, 2),
            'work_rate': round((work_days / 5) * 100, 1),
            'overtime_rate': round((overtime_minutes / weekly_standard_minutes) * 100, 1) if weekly_standard_minutes > 0 else 0,
            'average_daily_hours': round(actual_work_minutes / max(work_days, 1) / 60, 2),
        }
    
    def calculate_monthly_summary(self, user: User, year: int, month: int) -> Dict:
        """
        월간 근무시간 요약 계산
        
        Args:
            user: 대상 사용자
            year: 연도
            month: 월
            
        Returns:
            Dict: 월간 통계 데이터
        """
        start_date = date(year, month, 1)
        if month == 12:
            end_date = date(year + 1, 1, 1) - timedelta(days=1)
        else:
            end_date = date(year, month + 1, 1) - timedelta(days=1)
        
        records = WorkTimeRecord.objects.filter(
            user=user,
            work_date__range=[start_date, end_date]
        ).aggregate(
            total_work=Sum('total_work_minutes'),
            actual_work=Sum('actual_work_minutes'),
            overtime=Sum('overtime_minutes'),
            work_days=Count('id', filter=Q(actual_work_minutes__gt=0))
        )
        
        # 월간 예상 근무일수 계산 (주말 제외)
        expected_work_days = 0
        current_date = start_date
        while current_date <= end_date:
            if current_date.weekday() < 5:  # 월-금
                expected_work_days += 1
            current_date += timedelta(days=1)
        
        total_work_minutes = records['total_work'] or 0
        actual_work_minutes = records['actual_work'] or 0
        overtime_minutes = records['overtime'] or 0
        work_days = records['work_days'] or 0
        
        # 월간 표준 근무시간
        monthly_standard_minutes = expected_work_days * 480  # 일 8시간
        
        # 지각/조퇴/결근 통계
        standard_start = time(9, 0)
        standard_end = time(18, 0)
        
        late_count = WorkTimeRecord.objects.filter(
            user=user,
            work_date__range=[start_date, end_date],
            check_in_time__time__gt=standard_start
        ).count()
        
        early_leave_count = WorkTimeRecord.objects.filter(
            user=user,
            work_date__range=[start_date, end_date],
            check_out_time__time__lt=standard_end,
            actual_work_minutes__lt=480
        ).count()
        
        absent_days = expected_work_days - work_days
        
        return {
            'year': year,
            'month': month,
            'period': f"{year}-{month:02d}",
            'start_date': start_date,
            'end_date': end_date,
            'total_work_minutes': total_work_minutes,
            'actual_work_minutes': actual_work_minutes,
            'overtime_minutes': overtime_minutes,
            'work_days': work_days,
            'expected_work_days': expected_work_days,
            'absent_days': absent_days,
            'late_count': late_count,
            'early_leave_count': early_leave_count,
            'total_work_hours': round(total_work_minutes / 60, 2),
            'actual_work_hours': round(actual_work_minutes / 60, 2),
            'overtime_hours': round(overtime_minutes / 60, 2),
            'work_rate': round((work_days / expected_work_days) * 100, 1) if expected_work_days > 0 else 0,
            'attendance_rate': round(((expected_work_days - absent_days) / expected_work_days) * 100, 1) if expected_work_days > 0 else 0,
            'overtime_rate': round((overtime_minutes / monthly_standard_minutes) * 100, 1) if monthly_standard_minutes > 0 else 0,
            'average_daily_hours': round(actual_work_minutes / max(work_days, 1) / 60, 2),
        }
    
    def generate_work_time_chart_data(self, user: User, start_date: date, end_date: date) -> Dict:
        """
        근무시간 차트 데이터 생성 (PWA 차트용)
        
        Args:
            user: 대상 사용자
            start_date: 시작 날짜
            end_date: 종료 날짜
            
        Returns:
            Dict: 차트 데이터
        """
        records = WorkTimeRecord.objects.filter(
            user=user,
            work_date__range=[start_date, end_date]
        ).order_by('work_date')
        
        chart_data = {
            'labels': [],
            'datasets': {
                'actual_work_hours': [],
                'overtime_hours': [],
                'standard_line': []  # 8시간 기준선
            }
        }
        
        current_date = start_date
        while current_date <= end_date:
            chart_data['labels'].append(current_date.strftime('%m/%d'))
            
            record = records.filter(work_date=current_date).first()
            if record:
                chart_data['datasets']['actual_work_hours'].append(record.actual_work_hours)
                chart_data['datasets']['overtime_hours'].append(record.overtime_hours)
            else:
                chart_data['datasets']['actual_work_hours'].append(0)
                chart_data['datasets']['overtime_hours'].append(0)
            
            chart_data['datasets']['standard_line'].append(8.0)  # 8시간 기준
            current_date += timedelta(days=1)
        
        return chart_data
    
    def update_summary_cache(self, user: User, summary_type: str, year: int, 
                           month: int = None, week: int = None, day: int = None):
        """
        통계 요약 캐시 업데이트
        
        Args:
            user: 대상 사용자
            summary_type: 요약 타입 (daily, weekly, monthly, yearly)
            year: 연도
            month: 월 (선택적)
            week: 주 (선택적) 
            day: 일 (선택적)
        """
        try:
            with transaction.atomic():
                summary, created = WorkTimeSummary.objects.get_or_create(
                    user=user,
                    summary_type=summary_type,
                    year=year,
                    month=month,
                    week=week,
                    day=day
                )
                
                if summary_type == WorkTimeSummary.SummaryType.MONTHLY:
                    stats = self.calculate_monthly_summary(user, year, month)
                    
                    summary.total_work_minutes = stats['total_work_minutes']
                    summary.actual_work_minutes = stats['actual_work_minutes']
                    summary.overtime_minutes = stats['overtime_minutes']
                    summary.work_days = stats['work_days']
                    summary.expected_work_days = stats['expected_work_days']
                    summary.late_count = stats['late_count']
                    summary.early_leave_count = stats['early_leave_count']
                    summary.standard_work_minutes = stats['expected_work_days'] * 480
                
                elif summary_type == WorkTimeSummary.SummaryType.WEEKLY:
                    stats = self.calculate_weekly_summary(user, year, week)
                    
                    summary.total_work_minutes = stats['total_work_minutes']
                    summary.actual_work_minutes = stats['actual_work_minutes']
                    summary.overtime_minutes = stats['overtime_minutes']
                    summary.work_days = stats['work_days']
                    summary.expected_work_days = stats['expected_work_days']
                    summary.late_count = stats['late_count']
                    summary.early_leave_count = stats['early_leave_count']
                    summary.standard_work_minutes = 2400  # 주 40시간
                
                summary.save()
                
                logger.info(f"통계 캐시 업데이트 완료: {user.username} - {summary_type} {year}")
                
        except Exception as e:
            logger.error(f"통계 캐시 업데이트 실패: {user.username} - {str(e)}")
            raise


class ExcelExportService:
    """
    엑셀 내보내기 서비스
    """
    
    def export_monthly_report(self, user: User, year: int, month: int) -> bytes:
        """
        월간 근무시간 리포트 엑셀 내보내기
        
        Args:
            user: 대상 사용자
            year: 연도
            month: 월
            
        Returns:
            bytes: 엑셀 파일 바이너리 데이터
        """
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, PatternFill
            from openpyxl.utils import get_column_letter
            
            wb = Workbook()
            ws = wb.active
            ws.title = f"{year}년 {month}월 근무시간"
            
            # 헤더 설정
            headers = [
                '날짜', '요일', '출근시간', '퇴근시간', 
                '총 근무시간', '실 근무시간', '초과시간', 
                '상태', '메모'
            ]
            
            # 헤더 스타일
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_font = Font(color="FFFFFF", bold=True)
            
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col)
                cell.value = header
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center")
            
            # 데이터 입력
            start_date = date(year, month, 1)
            if month == 12:
                end_date = date(year + 1, 1, 1) - timedelta(days=1)
            else:
                end_date = date(year, month + 1, 1) - timedelta(days=1)
            
            records = WorkTimeRecord.objects.filter(
                user=user,
                work_date__range=[start_date, end_date]
            ).order_by('work_date')
            
            row = 2
            current_date = start_date
            
            while current_date <= end_date:
                record = records.filter(work_date=current_date).first()
                
                # 요일 계산
                weekday_names = ['월', '화', '수', '목', '금', '토', '일']
                weekday = weekday_names[current_date.weekday()]
                
                if record:
                    ws.cell(row=row, column=1).value = current_date.strftime('%Y-%m-%d')
                    ws.cell(row=row, column=2).value = weekday
                    ws.cell(row=row, column=3).value = record.check_in_time.strftime('%H:%M') if record.check_in_time else ''
                    ws.cell(row=row, column=4).value = record.check_out_time.strftime('%H:%M') if record.check_out_time else ''
                    ws.cell(row=row, column=5).value = record.get_work_time_formatted()
                    ws.cell(row=row, column=6).value = f"{record.actual_work_hours:.2f}h"
                    ws.cell(row=row, column=7).value = f"{record.overtime_hours:.2f}h" if record.overtime_hours > 0 else ''
                    ws.cell(row=row, column=8).value = record.get_status_display()
                    ws.cell(row=row, column=9).value = record.memo
                else:
                    ws.cell(row=row, column=1).value = current_date.strftime('%Y-%m-%d')
                    ws.cell(row=row, column=2).value = weekday
                    ws.cell(row=row, column=8).value = '결근' if current_date.weekday() < 5 else '휴무'
                
                # 주말 배경색
                if current_date.weekday() >= 5:
                    weekend_fill = PatternFill(start_color="F0F0F0", end_color="F0F0F0", fill_type="solid")
                    for col in range(1, len(headers) + 1):
                        ws.cell(row=row, column=col).fill = weekend_fill
                
                row += 1
                current_date += timedelta(days=1)
            
            # 요약 통계 추가
            stats_service = WorkTimeStatisticsService()
            monthly_stats = stats_service.calculate_monthly_summary(user, year, month)
            
            # 빈 행 추가
            row += 1
            
            # 통계 헤더
            ws.cell(row=row, column=1).value = "월간 통계"
            ws.cell(row=row, column=1).font = Font(bold=True)
            row += 1
            
            # 통계 데이터
            stats_data = [
                ['총 근무일수', f"{monthly_stats['work_days']}일"],
                ['예상 근무일수', f"{monthly_stats['expected_work_days']}일"],
                ['출근율', f"{monthly_stats['work_rate']:.1f}%"],
                ['총 근무시간', f"{monthly_stats['total_work_hours']:.2f}시간"],
                ['실 근무시간', f"{monthly_stats['actual_work_hours']:.2f}시간"],
                ['초과근무시간', f"{monthly_stats['overtime_hours']:.2f}시간"],
                ['지각 횟수', f"{monthly_stats['late_count']}회"],
                ['조퇴 횟수', f"{monthly_stats['early_leave_count']}회"],
            ]
            
            for stat_name, stat_value in stats_data:
                ws.cell(row=row, column=1).value = stat_name
                ws.cell(row=row, column=2).value = stat_value
                row += 1
            
            # 열 너비 조정
            for col in range(1, len(headers) + 1):
                ws.column_dimensions[get_column_letter(col)].width = 12
            
            # 바이너리 데이터로 변환
            from io import BytesIO
            buffer = BytesIO()
            wb.save(buffer)
            buffer.seek(0)
            
            return buffer.getvalue()
            
        except Exception as e:
            logger.error(f"엑셀 내보내기 실패: {user.username} - {year}-{month} - {str(e)}")
            raise