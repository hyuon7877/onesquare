"""
OneSquare 매출 관리 시스템 - 서비스 레이어
매출 통계, 리포트 생성, 분석 로직
"""

from django.db.models import Sum, Count, Avg, F, Q, Case, When, DecimalField, IntegerField
from django.utils import timezone
from datetime import datetime, date, timedelta
from decimal import Decimal
from collections import defaultdict
import pandas as pd
import json
import logging

from .models import (
    RevenueRecord, RevenueCategory, Client, Project, 
    RevenueTarget, RevenueReport
)
from .permissions import RevenuePermissionManager, UserRole

logger = logging.getLogger(__name__)

class RevenueAnalyticsService:
    """매출 분석 서비스"""
    
    @staticmethod
    def get_revenue_summary(user, start_date=None, end_date=None):
        """매출 요약 통계"""
        # 권한 기반 쿼리셋
        queryset = RevenuePermissionManager.filter_revenue_queryset(
            RevenueRecord.objects.filter(is_confirmed=True), user
        )
        
        # 기간 필터
        if start_date:
            queryset = queryset.filter(revenue_date__gte=start_date)
        if end_date:
            queryset = queryset.filter(revenue_date__lte=end_date)
        
        # 기본 통계
        summary = queryset.aggregate(
            total_revenue=Sum('net_amount'),
            total_count=Count('id'),
            avg_revenue=Avg('net_amount')
        )
        
        # 결제 상태별 통계
        payment_stats = queryset.values('payment_status').annotate(
            count=Count('id'),
            amount=Sum('net_amount')
        )
        
        # 카테고리별 통계
        category_stats = queryset.values(
            'category__name', 'category__code'
        ).annotate(
            count=Count('id'),
            amount=Sum('net_amount')
        ).order_by('-amount')
        
        return {
            'summary': summary,
            'payment_stats': list(payment_stats),
            'category_stats': list(category_stats),
            'period': {
                'start_date': start_date.isoformat() if start_date else None,
                'end_date': end_date.isoformat() if end_date else None
            }
        }
    
    @staticmethod
    def get_monthly_trend(user, months=12):
        """월별 매출 트렌드"""
        queryset = RevenuePermissionManager.filter_revenue_queryset(
            RevenueRecord.objects.filter(is_confirmed=True), user
        )
        
        monthly_data = []
        for i in range(months):
            # 월 시작/종료일 계산
            target_date = date.today() - timedelta(days=30*i)
            month_start = target_date.replace(day=1)
            
            if month_start.month == 12:
                month_end = month_start.replace(year=month_start.year+1, month=1)
            else:
                month_end = month_start.replace(month=month_start.month+1)
            
            # 월별 통계
            month_stats = queryset.filter(
                revenue_date__gte=month_start,
                revenue_date__lt=month_end
            ).aggregate(
                total_revenue=Sum('net_amount'),
                total_count=Count('id'),
                confirmed_revenue=Sum(
                    Case(
                        When(payment_status='completed', then=F('net_amount')),
                        default=0,
                        output_field=DecimalField()
                    )
                )
            )
            
            monthly_data.append({
                'month': month_start.strftime('%Y-%m'),
                'year': month_start.year,
                'month_num': month_start.month,
                'total_revenue': float(month_stats['total_revenue'] or 0),
                'confirmed_revenue': float(month_stats['confirmed_revenue'] or 0),
                'count': month_stats['total_count'] or 0
            })
        
        return monthly_data[::-1]  # 오래된 순서로 정렬
    
    @staticmethod
    def get_client_analysis(user, limit=20):
        """고객별 매출 분석"""
        queryset = RevenuePermissionManager.filter_revenue_queryset(
            RevenueRecord.objects.filter(is_confirmed=True), user
        )
        
        client_stats = queryset.values(
            'client__id', 'client__name', 'client__code', 'client__client_type'
        ).annotate(
            total_revenue=Sum('net_amount'),
            avg_revenue=Avg('net_amount'),
            count=Count('id'),
            last_revenue_date=Max('revenue_date'),
            confirmed_count=Count(
                Case(
                    When(payment_status='completed', then=1),
                    output_field=IntegerField()
                )
            )
        ).order_by('-total_revenue')[:limit]
        
        # 추가 분석
        for client in client_stats:
            # 완료율 계산
            if client['count'] > 0:
                client['completion_rate'] = (client['confirmed_count'] / client['count']) * 100
            else:
                client['completion_rate'] = 0
                
            # 최근 활동 (일수)
            if client['last_revenue_date']:
                days_since_last = (date.today() - client['last_revenue_date']).days
                client['days_since_last'] = days_since_last
            else:
                client['days_since_last'] = None
        
        return list(client_stats)
    
    @staticmethod
    def get_project_analysis(user, limit=20):
        """프로젝트별 매출 분석"""
        queryset = RevenuePermissionManager.filter_revenue_queryset(
            RevenueRecord.objects.filter(is_confirmed=True), user
        )
        
        project_stats = queryset.values(
            'project__id', 'project__name', 'project__code', 
            'project__status', 'project__contract_amount'
        ).annotate(
            total_revenue=Sum('net_amount'),
            avg_revenue=Avg('net_amount'),
            count=Count('id'),
            last_revenue_date=Max('revenue_date')
        ).order_by('-total_revenue')[:limit]
        
        # 프로젝트 완료율 계산
        for project in project_stats:
            contract_amount = project.get('project__contract_amount', 0)
            if contract_amount and contract_amount > 0:
                completion_rate = (project['total_revenue'] / float(contract_amount)) * 100
                project['completion_rate'] = min(completion_rate, 100)  # 100% 초과 방지
            else:
                project['completion_rate'] = 0
        
        return list(project_stats)
    
    @staticmethod
    def get_revenue_forecast(user, months_ahead=6):
        """매출 예측 (간단한 트렌드 기반)"""
        # 지난 12개월 데이터로 트렌드 분석
        historical_data = RevenueAnalyticsService.get_monthly_trend(user, 12)
        
        if len(historical_data) < 3:
            return []  # 데이터가 부족하면 예측 불가
        
        # 최근 6개월 평균 성장률 계산
        recent_months = historical_data[-6:]
        growth_rates = []
        
        for i in range(1, len(recent_months)):
            prev_revenue = recent_months[i-1]['total_revenue']
            curr_revenue = recent_months[i]['total_revenue']
            
            if prev_revenue > 0:
                growth_rate = (curr_revenue - prev_revenue) / prev_revenue
                growth_rates.append(growth_rate)
        
        # 평균 성장률
        avg_growth_rate = sum(growth_rates) / len(growth_rates) if growth_rates else 0
        
        # 예측 데이터 생성
        forecast_data = []
        last_revenue = historical_data[-1]['total_revenue']
        
        for i in range(1, months_ahead + 1):
            # 다음 달 계산
            next_month = date.today() + timedelta(days=30*i)
            
            # 트렌드 기반 예측 (단순화)
            predicted_revenue = last_revenue * (1 + avg_growth_rate) ** i
            
            forecast_data.append({
                'month': next_month.strftime('%Y-%m'),
                'year': next_month.year,
                'month_num': next_month.month,
                'predicted_revenue': max(0, predicted_revenue),  # 음수 방지
                'confidence': max(0.1, 0.8 - (i * 0.1))  # 시간이 지날수록 신뢰도 감소
            })
        
        return forecast_data

class RevenueReportService:
    """매출 리포트 생성 서비스"""
    
    @staticmethod
    def generate_monthly_report(user, year, month):
        """월간 매출 리포트 생성"""
        start_date = date(year, month, 1)
        if month == 12:
            end_date = date(year + 1, 1, 1)
        else:
            end_date = date(year, month + 1, 1)
        
        # 기본 통계
        summary = RevenueAnalyticsService.get_revenue_summary(
            user, start_date, end_date
        )
        
        # 상세 분석
        client_analysis = RevenueAnalyticsService.get_client_analysis(user, 10)
        project_analysis = RevenueAnalyticsService.get_project_analysis(user, 10)
        
        # 목표 대비 달성률
        targets = RevenueTarget.objects.filter(
            target_type='monthly',
            year=year,
            month=month
        )
        
        target_analysis = []
        for target in targets:
            achievement_rate = target.get_achievement_rate()
            target_analysis.append({
                'id': str(target.id),
                'description': str(target),
                'target_amount': float(target.target_amount),
                'achievement_rate': achievement_rate,
                'category': target.category.name if target.category else '전체'
            })
        
        # 리포트 데이터 구성
        report_data = {
            'period': {
                'year': year,
                'month': month,
                'start_date': start_date.isoformat(),
                'end_date': (end_date - timedelta(days=1)).isoformat()
            },
            'summary': summary,
            'client_analysis': client_analysis,
            'project_analysis': project_analysis,
            'target_analysis': target_analysis,
            'generated_at': timezone.now().isoformat(),
            'generated_by': user.get_full_name() or user.username
        }
        
        # 권한에 따른 데이터 마스킹
        user_role = RevenuePermissionManager.get_user_role(user)
        if user_role not in [UserRole.SUPER_ADMIN, UserRole.ADMIN]:
            report_data = RevenuePermissionManager.mask_revenue_data(report_data, user)
        
        return report_data
    
    @staticmethod
    def generate_quarterly_report(user, year, quarter):
        """분기별 매출 리포트 생성"""
        quarter_months = {
            1: (1, 3),  # Q1: 1-3월
            2: (4, 6),  # Q2: 4-6월
            3: (7, 9),  # Q3: 7-9월
            4: (10, 12) # Q4: 10-12월
        }
        
        start_month, end_month = quarter_months[quarter]
        start_date = date(year, start_month, 1)
        
        if end_month == 12:
            end_date = date(year + 1, 1, 1)
        else:
            end_date = date(year, end_month + 1, 1)
        
        # 분기별 상세 분석
        summary = RevenueAnalyticsService.get_revenue_summary(
            user, start_date, end_date
        )
        
        # 월별 세부 데이터
        monthly_breakdown = []
        for month in range(start_month, end_month + 1):
            month_data = RevenueAnalyticsService.get_revenue_summary(
                user, 
                date(year, month, 1),
                date(year, month + 1, 1) if month < 12 else date(year + 1, 1, 1)
            )
            monthly_breakdown.append({
                'month': month,
                'month_name': date(year, month, 1).strftime('%B'),
                'data': month_data
            })
        
        # 분기 목표 달성률
        quarterly_targets = RevenueTarget.objects.filter(
            target_type='quarterly',
            year=year,
            quarter=quarter
        )
        
        target_analysis = []
        for target in quarterly_targets:
            achievement_rate = target.get_achievement_rate()
            target_analysis.append({
                'id': str(target.id),
                'description': str(target),
                'target_amount': float(target.target_amount),
                'achievement_rate': achievement_rate
            })
        
        report_data = {
            'period': {
                'year': year,
                'quarter': quarter,
                'start_date': start_date.isoformat(),
                'end_date': (end_date - timedelta(days=1)).isoformat()
            },
            'summary': summary,
            'monthly_breakdown': monthly_breakdown,
            'target_analysis': target_analysis,
            'generated_at': timezone.now().isoformat(),
            'generated_by': user.get_full_name() or user.username
        }
        
        return report_data
    
    @staticmethod
    def generate_excel_report(report_data, report_type='monthly'):
        """엑셀 리포트 생성"""
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Border, Side
            from openpyxl.utils.dataframe import dataframe_to_rows
            
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = f"매출리포트_{report_type}"
            
            # 헤더 스타일
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            
            # 리포트 제목
            ws['A1'] = f"매출 {report_type} 리포트"
            ws['A1'].font = Font(bold=True, size=16)
            ws.merge_cells('A1:D1')
            
            # 기간 정보
            row = 3
            if 'period' in report_data:
                period = report_data['period']
                ws[f'A{row}'] = "리포트 기간:"
                ws[f'B{row}'] = f"{period.get('start_date')} ~ {period.get('end_date')}"
                row += 2
            
            # 요약 통계
            if 'summary' in report_data:
                summary = report_data['summary']['summary']
                ws[f'A{row}'] = "요약 통계"
                ws[f'A{row}'].font = header_font
                ws[f'A{row}'].fill = header_fill
                row += 1
                
                ws[f'A{row}'] = "총 매출"
                ws[f'B{row}'] = float(summary.get('total_revenue', 0))
                row += 1
                
                ws[f'A{row}'] = "매출 건수"
                ws[f'B{row}'] = summary.get('total_count', 0)
                row += 1
                
                ws[f'A{row}'] = "평균 매출"
                ws[f'B{row}'] = float(summary.get('avg_revenue', 0))
                row += 2
            
            # 고객별 분석
            if 'client_analysis' in report_data:
                ws[f'A{row}'] = "고객별 매출 분석"
                ws[f'A{row}'].font = header_font
                ws[f'A{row}'].fill = header_fill
                row += 1
                
                # 헤더
                headers = ['고객명', '총매출', '건수', '평균매출', '완료율']
                for col, header in enumerate(headers, 1):
                    cell = ws.cell(row=row, column=col, value=header)
                    cell.font = header_font
                    cell.fill = header_fill
                row += 1
                
                # 데이터
                for client in report_data['client_analysis']:
                    ws[f'A{row}'] = client.get('client__name', '')
                    ws[f'B{row}'] = float(client.get('total_revenue', 0))
                    ws[f'C{row}'] = client.get('count', 0)
                    ws[f'D{row}'] = float(client.get('avg_revenue', 0))
                    ws[f'E{row}'] = f"{client.get('completion_rate', 0):.1f}%"
                    row += 1
            
            # 컬럼 너비 자동 조정
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            return wb
            
        except ImportError:
            logger.error("openpyxl 라이브러리가 필요합니다.")
            return None
        except Exception as e:
            logger.error(f"엑셀 리포트 생성 실패: {e}")
            return None

class RevenueTargetService:
    """매출 목표 관리 서비스"""
    
    @staticmethod
    def check_target_achievements(user):
        """매출 목표 달성 현황 확인"""
        current_date = date.today()
        current_year = current_date.year
        current_month = current_date.month
        current_quarter = (current_month - 1) // 3 + 1
        
        # 사용자 권한에 따른 목표 필터링
        targets = RevenueTarget.objects.filter(year=current_year)
        user_role = RevenuePermissionManager.get_user_role(user)
        
        if user_role == UserRole.TEAM_MEMBER:
            targets = targets.filter(assigned_user=user)
        elif user_role == UserRole.MIDDLE_MANAGER:
            targets = targets.filter(
                Q(assigned_user=user) | 
                Q(assigned_user__isnull=True)
            )
        
        achievements = []
        for target in targets:
            achievement_rate = target.get_achievement_rate()
            
            # 달성 상태 결정
            if achievement_rate >= 100:
                status = 'achieved'
            elif achievement_rate >= 80:
                status = 'on_track'
            elif achievement_rate >= 50:
                status = 'behind'
            else:
                status = 'critical'
            
            achievements.append({
                'target': target,
                'achievement_rate': achievement_rate,
                'status': status,
                'needs_attention': achievement_rate < 80
            })
        
        return achievements
    
    @staticmethod
    def get_target_alerts(user):
        """목표 달성 관련 알림"""
        achievements = RevenueTargetService.check_target_achievements(user)
        alerts = []
        
        for achievement in achievements:
            target = achievement['target']
            rate = achievement['achievement_rate']
            
            # 임계치별 알림 생성
            if rate < 50:
                alerts.append({
                    'level': 'critical',
                    'title': f'{target} 목표 달성률 위험',
                    'message': f'현재 달성률: {rate:.1f}% - 즉시 조치가 필요합니다.',
                    'target_id': target.id
                })
            elif rate < 80:
                alerts.append({
                    'level': 'warning', 
                    'title': f'{target} 목표 달성률 저조',
                    'message': f'현재 달성률: {rate:.1f}% - 추가 노력이 필요합니다.',
                    'target_id': target.id
                })
            elif rate >= 100:
                alerts.append({
                    'level': 'success',
                    'title': f'{target} 목표 달성 완료',
                    'message': f'달성률: {rate:.1f}% - 축하합니다!',
                    'target_id': target.id
                })
        
        return alerts

class RevenueComparisonService:
    """매출 비교 분석 서비스"""
    
    @staticmethod
    def compare_periods(user, period1_start, period1_end, period2_start, period2_end):
        """기간별 매출 비교"""
        period1_summary = RevenueAnalyticsService.get_revenue_summary(
            user, period1_start, period1_end
        )
        period2_summary = RevenueAnalyticsService.get_revenue_summary(
            user, period2_start, period2_end
        )
        
        # 증감률 계산
        p1_revenue = float(period1_summary['summary']['total_revenue'] or 0)
        p2_revenue = float(period2_summary['summary']['total_revenue'] or 0)
        
        if p2_revenue > 0:
            growth_rate = ((p1_revenue - p2_revenue) / p2_revenue) * 100
        else:
            growth_rate = 0
        
        return {
            'period1': {
                'start_date': period1_start.isoformat(),
                'end_date': period1_end.isoformat(),
                'data': period1_summary
            },
            'period2': {
                'start_date': period2_start.isoformat(), 
                'end_date': period2_end.isoformat(),
                'data': period2_summary
            },
            'comparison': {
                'revenue_growth_rate': growth_rate,
                'revenue_difference': p1_revenue - p2_revenue,
                'is_improving': growth_rate > 0
            }
        }
    
    @staticmethod
    def year_over_year_comparison(user, current_year=None):
        """전년 대비 매출 비교"""
        if not current_year:
            current_year = date.today().year
            
        previous_year = current_year - 1
        
        current_year_data = RevenueAnalyticsService.get_revenue_summary(
            user,
            date(current_year, 1, 1),
            date(current_year + 1, 1, 1)
        )
        
        previous_year_data = RevenueAnalyticsService.get_revenue_summary(
            user,
            date(previous_year, 1, 1),
            date(previous_year + 1, 1, 1)
        )
        
        return RevenueComparisonService.compare_periods(
            user,
            date(current_year, 1, 1), date(current_year + 1, 1, 1),
            date(previous_year, 1, 1), date(previous_year + 1, 1, 1)
        )